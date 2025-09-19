import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ───── Streamlit setup ────────────────────────────────────────────────────────
st.set_page_config(page_title="Visitor List Cleaner (US)", layout="wide")
st.title("🇺🇸 Clarity Gate - US Visitor Data Cleaning & Validation 🫧")

st.info(
    """
    **Data Integrity Is Our Foundation**  
    At every step—from file upload to final report—we enforce strict validation to guarantee your visitor data is accurate, complete, and compliant.  
    Maintaining integrity not only expedites gate clearance, it protects our facilities and ensures we meet all regulatory requirements.  

    **Why is Data Integrity Important?**  
    **Accuracy**: Correct visitor details reduce clearance delays.  
    **Security**: Reliable ID checks prevent unauthorized access.  
    **Compliance**: Audit-ready records ensure regulatory adherence.  
    **Efficiency**: Trustworthy data powers faster reporting and analytics.
    """
)

# ───── 3) Uploader & Warning ───────────────────────────────────────────────────

st.markdown(
    """<div style='font-size:16px; font-weight:bold; color:#38761d;'>
    Please ensure your spreadsheet has no missing or malformed fields.<br>
    Columns E (First Name) and Column F (Middle and Last Name) are not required to be filled in.<br>
    </div>""",
    unsafe_allow_html=True
)

#uploaded = st.file_uploader("📁 Upload file", type=["xlsx"])

# ───── Download Sample Template ───────────────────────────────────────────────
with open("US_Template.xlsx", "rb") as f:
    st.download_button(
        label="⭐️ Download US Template",
        data=f,
        file_name="US_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ───── 4) Estimate Clearance Date ───────────────────────────────────────────────
us_tz = ZoneInfo("America/New_York")
now = datetime.now(us_tz)
formatted_now = now.strftime("%A %d %B, %I:%M%p").lstrip("0")
st.write("**Today (US/Eastern Time):**", formatted_now)

def next_working_day(d):
    """Return the next calendar date that is a weekday (Mon–Fri)."""
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def earliest_clearance_inclusive(submit_dt: datetime, workdays: int = 2) -> datetime:
    """
    Business rule:
    - Start counting working days from the submission calendar day itself (inclusive).
    - If submission falls on a weekend, start from the next Monday.
    - After counting N working days, the earliest clearance is the **next** working day.
    """
    # Day 1 is the submission date (roll forward if weekend)
    d = next_working_day(submit_dt.date())

    # Count remaining working days (we already counted Day 1)
    counted = 1
    while counted < workdays:
        d += timedelta(days=1)
        if d.weekday() < 5:
            counted += 1

    # Earliest clearance is the NEXT working day after the Nth working day
    clearance = d + timedelta(days=1)
    clearance = next_working_day(clearance)
    return clearance

if st.button("▶️ Earliest clearance (US):"):
    clearance_date = earliest_clearance_inclusive(now, workdays=2)
    st.success(f" **{clearance_date:%A} {clearance_date.day} {clearance_date:%B}**")


# ───── Helper functions ────────────────────────────────────────────────────────

def split_name(full_name):
    s = str(full_name).strip()
    if " " in s:
        i = s.find(" ")
        return pd.Series([s[:i], s[i+1:]])
    return pd.Series([s, ""])

def clean_gender(g):
    v = str(g).strip().upper()
    if v == "M":
        return "Male"
    if v == "F":
        return "Female"
    if v in ("MALE","FEMALE"):
        return v.title()
    return v.title()

def fix_mobile(x):
    d = re.sub(r"\D", "", str(x))
    # if too long...
    if len(d) > 10:
        extra = len(d) - 10
        if d.endswith("0" * extra):
            d = d[:-extra]
        else:
            d = d[-10:]
    if len(d) < 10:
        d = d.zfill(10)
    return d

# ───── Core Cleaning Logic ────────────────────────────────────────────────────

def clean_data_us(df: pd.DataFrame) -> pd.DataFrame:
    # 1) Trim to exactly 11 cols then rename
    df = df.iloc[:, :11]
    df.columns = [
        "S/N",
        "Vehicle Plate Number",
        "Company Full Name",
        "Full Name",
        "First Name",
        "Middle and Last Name",
        "Driver License Number",
        "Nationality (Country Name)",
        "Gender",
        "Mobile Number",
        "Remarks", 
    ]

    # 2) Drop rows where all of Full Name → Mobile are blank
    df = df.dropna(subset=df.columns[3:10], how="all")

    # 3) Normalize nationality (incl. Indian → India, etc.)
    nat_map = {
        "chinese":     "China",
        "singaporean": "Singapore",
        "malaysian":   "Malaysia",
        "indian":      "India",
        "usa":         "United States",
        "us":          "United States",
    }
    df["Nationality (Country Name)"] = (
        df["Nationality (Country Name)"]
          .astype(str)
          .str.strip()
          .str.lower()
          .replace(nat_map, regex=False)
          .str.title()
    )

    # 4) Sort by Company → Country → Full Name
    df = df.sort_values(
        ["Company Full Name", "Nationality (Country Name)", "Full Name"],
        ignore_index=True,
    )

    # 5) Reset S/N
    df["S/N"] = range(1, len(df) + 1)

    # 6) Standardize Vehicle Plate Number
    df["Vehicle Plate Number"] = (
        df["Vehicle Plate Number"]
          .astype(str)
          .str.replace(r"[\/,]", ";", regex=True)
          .str.replace(r"\s*;\s*", ";", regex=True)
          .str.strip()
          .replace("nan","", regex=False)
    )

    # 7) Proper-case & split names (in case the template didn't pre-split)
    df["Full Name"] = df["Full Name"].astype(str).str.title()
    df[["First Name","Middle and Last Name"]] = (
        df["Full Name"].apply(split_name)
    )

    # 8) Fix Mobile Number → 10 digits
    df["Mobile Number"] = df["Mobile Number"].apply(fix_mobile)

    # 9) Normalize gender
    df["Gender"] = df["Gender"].apply(clean_gender)


    # 10) Driver License Number: remove spaces, keep last 4 chars
    df["Driver License Number"] = (
    df["Driver License Number"]
      .fillna("")                             # keep NaN as blank
      .astype(str)
      .str.replace(r"\s+", "", regex=True)    # remove ALL spaces
      .str[-4:]                               # last 4 characters
)

    return df

# ───── Build & style the single-sheet Excel ──────────────────────────────────

def generate_visitor_only_us(df: pd.DataFrame) -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Visitor List")
        ws = writer.sheets["Visitor List"]

        # styling objects
        header_fill  = PatternFill("solid", fgColor="94B455")
        border       = Border(Side("thin"),Side("thin"),Side("thin"),Side("thin"))
        center       = Alignment("center","center")
        normal_font  = Font(name="Calibri", size=9)
        bold_font    = Font(name="Calibri", size=9, bold=True)

        # 1) Apply borders, alignment, font
        for row in ws.iter_rows():
            for cell in row:
                cell.border    = border
                cell.alignment = center
                cell.font      = normal_font

        # 2) Style header row
        for col in range(1, ws.max_column + 1):
            h = ws[f"{get_column_letter(col)}1"]
            h.fill = header_fill
            h.font = bold_font

        # 3) Freeze top row
        ws.freeze_panes = ws["A2"]

        # 4) Auto-fit columns & set row height
        for col in ws.columns:
            w = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[get_column_letter(col[0].column)].width = w + 2
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 20

        # 5) Vehicles summary
        plates = []
        for v in df["Vehicle Plate Number"].dropna():
            plates += [x.strip() for x in str(v).split(";") if x.strip()]
        ins = ws.max_row + 2
        if plates:
            ws[f"B{ins}"].value     = "Vehicles"
            ws[f"B{ins}"].border    = border
            ws[f"B{ins}"].alignment = center
            ws[f"B{ins+1}"].value   = ";".join(sorted(set(plates)))
            ws[f"B{ins+1}"].border  = border
            ws[f"B{ins+1}"].alignment = center
            ins += 2

        # 6) Total Visitors
        ws[f"B{ins}"].value     = "Total Visitors"
        ws[f"B{ins}"].border    = border
        ws[f"B{ins}"].alignment = center
        ws[f"B{ins+1}"].value   = df["Company Full Name"].notna().sum()
        ws[f"B{ins+1}"].border  = border
        ws[f"B{ins+1}"].alignment = center

    buf.seek(0)
    return buf

# ───── Streamlit UI: Upload & Download ────────────────────────────────────────
uploaded = st.file_uploader("📁 Upload File", type=["xlsx"])
if uploaded:
    raw_df = pd.read_excel(uploaded, sheet_name="Visitor List")
    cleaned = clean_data_us(raw_df)
    out_buf = generate_visitor_only_us(cleaned)

    # Build filename: CompanyName_YYYYMMDD.xlsx in US/Eastern time
    today = datetime.now(ZoneInfo("America/New_York")).strftime("%Y%m%d")
    company_cell = raw_df.iloc[0, 2]
    company = (
        str(company_cell).strip()
        if pd.notna(company_cell) and str(company_cell).strip()
        else "VisitorList"
    )
    fname = f"{company}_{today}.xlsx"

    st.download_button(
        label="📥 Download Cleaned Visitor List (US)",
        data=out_buf.getvalue(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ───── 5) Final Notice (always shown) ────────────────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:16px;">
      We will do our utmost to deliver your access ticket 1 day before your scheduled entry.<br>
      Kindly ensure that approved access clearance codes are obtained before planning or commencing any work activities in the data center.<br>
      Please be reminded to go through the Clarity Gate prior to submission, and ensure that all visitor and shipment details are complete and accurate to prevent rescheduling due to clarification.<br><br>
      <strong>Note:</strong><br>
      The Clarity Gate operates on the GOFAI system, which relies on explicitly programmed rules and logic.<br>
      Although its validation accuracy can reach up to 98%, we strongly recommend that you thoroughly review all information before submission.<br>
      Thank you for your cooperation.<br><br>
    </div>
    """,
    unsafe_allow_html=True,
)


# ───── 6) Vendor Accuracy Reminder (always shown) ────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:16px;">
      <strong>Kindly remind all vendors to take the accuracy of the submitted information seriously.</strong><br>
      Any <em>incorrect or incomplete details</em> will result in <em>rejection</em>, and the personnel will not be allowed to enter the data centre.<br>
      <em>This requirement is non-negotiable, and strict compliance is expected.<em><br>
      Please ensure this message is clearly conveyed to all concerned.
    </div>
    """,
    unsafe_allow_html=True,
)
